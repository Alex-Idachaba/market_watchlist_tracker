import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import sqlite3
import json
import os
import time
from datetime import datetime, timezone
from pathlib import Path
import shutil

load_dotenv()
API_KEY = os.getenv("API_KEY")

symbols = ["AUD/USD", "EUR/USD", "EUR/JPY", "GBP/USD", "USD/JPY", "USD/CAD", 
           "BTC/USD"]

today_date = datetime.now(timezone.utc)
raw_data_file_name = f"{str(today_date.date())}_raw.json"

base_path = Path.cwd()
raw_data_path = base_path / "data" / "raw" / raw_data_file_name
db_path = base_path / "data" / "database"
reports_folder = base_path / "reports"

def fetch_price_data(symbol, api_key):
    """
    Fetch daily time series data for a single instrument from Twelve Data.
    Returns the raw JSON response as a dict, or None on failure.
    """
    BASE_URL = (
        "https://api.twelvedata.com/time_series"
        "?apikey="
        + api_key
        + "&symbol="
        + symbol
        + "&interval=1day"
        + "&outputsize=1"
        + "&timezone=America/New_York"
        + "&format=JSON"
    )
    try:
        response = requests.get(BASE_URL, timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.ConnectionError:
        print("Network error - could not reach the Twelvedata API")
        return None
    except requests.exceptions.Timeout:
        print("Request timed out - the API took too long to respond")
        return None
    except requests.exceptions.HTTPError as e:
        print(f"API returned an error: {e}")
        return None

def fetch_watchlist_data(symbols, api_key):
    """
    Loop through a list of instrument symbols and fetch data for each.
    Returns a dict mapping symbol -> raw JSON response (or None if that symbol failed).
    """
    watchlist_data = {}

    for symbol in symbols:
        response = fetch_price_data(symbol, api_key)
        if not response:
            print(f"Could not fetch: {symbol} data")
            continue
        watchlist_data[symbol] = response
        time.sleep(8)

    return watchlist_data

def save_raw_json(watchlist_data, raw_data_file_name, raw_data_path):
    """
    Write raw API response data to a JSON file at the given path.
    Returns True if the write succeeded, False otherwise.
    """
    raw_folder_path = Path.cwd() / "data" / "raw"
    try:
        raw_folder_path.mkdir(parents=True, exist_ok=True)
        with open(raw_data_path, "w") as file:
            json.dump(watchlist_data, file)
        print(f"{raw_data_file_name} written to {raw_data_path}")
        return True
    
    except OSError as e:
        print(f"Failed to write Json file: {e}")
        return False

def load_raw_json(filepath):
    """
    Read a cached JSON file from disk and return it as a dict.
    """    
    try:
        with open(filepath, "r") as file:
            return json.load(file)
    except OSError as e:
        print(f"The specific file does not exist {e}")
        return None
    except json.JSONDecodeError as e:
        print(f"The file exists but isn't valid JSON: {e}")
        return None

def parse_watchlist_data(json_file_content):
    """
    Convert raw per-symbol JSON responses into a flat list of dictionaries,
    one per instrument per day.
    """
    symbols_data_list = []

    for symbol in json_file_content:
        symbol_data = json_file_content[symbol]

        if symbol_data is None:
            print(f"Skipping {symbol}: no data was fetched for this symbol")
            continue
        try:
            trades_data = {
                "instrument": symbol,
                "date": symbol_data["values"][0]["datetime"],
                "open": float(symbol_data["values"][0]["open"]),
                "high": float(symbol_data["values"][0]["high"]),
                "low": float(symbol_data["values"][0]["low"]),
                "close": float(symbol_data["values"][0]["close"]),
            }
            symbols_data_list.append(trades_data)
        except KeyError as e:
            print(f"Json file content is missing an expected field: {e}")
            continue

    return symbols_data_list

def build_date_folder(base_path, date):
    """
    Construct (and create if needed) a year/month folder path under base_path
    for the given date. Returns the folder as a Path object.
    """
    year = date.strftime("%Y")
    month = date.strftime("%m")
    folder_path = base_path / "data" / year / month

    try:
        folder_path.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        print(f"Failed to create folder: {e}")
        return None

    return folder_path

def archive_daily_files(source_paths, destination_folder):
    """
    Move or copy the given files into the destination folder.
    """
    if destination_folder is None:
        print("Cannot archive files: destination folder was not created")
        return None

    for source_path in source_paths:
        destination_path = destination_folder / source_path.name

        try:
            if destination_path.exists():
                destination_path.unlink()
            shutil.move(source_path, destination_folder)
        except OSError as e:
            print(f"Could not move {source_path} to destination folder: {e}")
            continue

    return destination_folder

def init_database(db_path):
    """
    Create the SQLite database and daily_prices table if they don't already exist.
    """
    db_file = "watchlist_history.db"
    connection = None

    try:
        db_path.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        print(f"Failed to create folder: {e}")
        return None
    try:
        connection = sqlite3.connect(db_path / db_file)
        cursor = connection.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS daily_prices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instrument TEXT NOT NULL,
                date TEXT NOT NULL,
                open REAL,
                high REAL,
                low REAL,
                close REAL    
            )
        """)
        connection.commit()
        return db_path / db_file
    
    except sqlite3.Error as e:
        print(f"Failed to setup the database: {e}")
        return None
    
    finally:
         if connection is not None:
             connection.close()

def record_exists(db_path, instrument="", date=""):
    """
    Check whether a row for this instrument and date already exists.
    Returns True or False.
    """
    connection = None

    try:
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM daily_prices WHERE instrument = ? AND date = ?",
                        (instrument, date))
        result = cursor.fetchone()

        if result is not None:
            return True
        else:
            return False
        
    except sqlite3.Error as e:
        print(f"Failed to check the database{e}")
        return False
    
    finally:
        if connection is not None:
            connection.close()
    
def insert_daily_records(db_path, records):
    """
    Insert a list of parsed records into the daily_prices table,
    skipping any that already exist for that instrument/date.
    """
    connection = None

    try:

        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()

        for row in records:
            instrument = row["instrument"]
            date = row["date"]
            open_price = float(row["open"])
            high = float(row["high"])
            low = float(row["low"])
            close = float(row["close"])
            
            already_exists = record_exists(db_path, instrument=instrument, date=date)

            if already_exists is not True:
                cursor.execute(
                    "INSERT INTO daily_prices (instrument, date, open, high, low, close) \
                    VALUES (?, ?, ?, ?, ?, ?)",
                    (instrument, date, open_price, high, low, close)
                )
        connection.commit()
        print("Daily Prices successfully inserted into database.")
        return True
    
    except sqlite3.Error as e:
        print(f"Failed to insert values into database: {e}")
        return False
    
    finally:
        if connection is not None:
            connection.close()

def fetch_last_n_days(db_path, n_days):
    """
    Query the daily_prices table for the last n_days of data.
    Returns a list of records.
    """
    connection = None

    try:
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()

        cursor.execute(
            """
            SELECT * 
            FROM daily_prices
            WHERE date >= date('now', '-' || ? || ' days')
            ORDER BY date DESC;
            """,
            (n_days,)
        )
        rows = cursor.fetchall()
        return rows
    
    except sqlite3.Error as e:
        print(f"Failed to retrieve data: {e}")
        return None

    finally:
        if connection is not None:
            connection.close()

def build_weekly_report(records, output_path):
    """
    Build and save a formatted Excel workbook summarizing the given records.
    """
    instruments = ["AUD/USD", "EUR/USD", "EUR/JPY", "GBP/USD", "USD/JPY", "USD/CAD",
                    "BTC/USD"]

    grid_data = {}
    range_data = {}

    for row in records:
        instrument = row[1]
        date = row[2]
        high = row[4]
        low = row[5]
        close = row[6]

        if date not in grid_data:
            grid_data[date] = {}
        grid_data[date][instrument] = close

        if date not in range_data:
            range_data[date] = {}
        range_data[date][instrument] = (high - low) / close

    dates = sorted(grid_data.keys())

    start_date = dates[0]
    end_date_full = dates[-1]
    end_date_short = end_date_full[5:]
    filename = f"week_{start_date}_to_{end_date_short}.xlsx"
    full_output_path = output_path / filename

    try:
        workbook = Workbook()
        sheet = workbook.active

        bold_font = Font(bold=True)
        large_range_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                        fill_type="solid")

        sheet.cell(row=1, column=1, value="Date").font = bold_font

        for col_index, instrument in enumerate(instruments, start=2):
            sheet.cell(row=1, column=col_index, value=instrument).font = bold_font

        for row_index, date in enumerate(dates, start=2):
            sheet.cell(row=row_index, column=1, value=date).font = bold_font

            for col_index, instrument in enumerate(instruments, start=2):
                day_data = grid_data.get(date, {})
                close_price = day_data.get(instrument, None)
                cell = sheet.cell(row=row_index, column=col_index, 
                                  value=close_price)

                day_range_data = range_data.get(date, {})
                range_percent = day_range_data.get(instrument, None)

                if range_percent is not None:
                    threshold = 0.03 if "BTC" in instrument else 0.01
                    if range_percent > threshold:
                        cell.fill = large_range_fill

        sheet.column_dimensions["A"].width = 12
        for col_index in range(2, len(instruments) + 2):
            column_letter = sheet.cell(row=1, column=col_index).column_letter
            sheet.column_dimensions[column_letter].width = 12

        workbook.save(full_output_path)
        print(f"Weekly report saved to {full_output_path}")
        return True

    except OSError as e:
        print(f"Failed to save the report: {e}")
        return False

def main():
    
    watchlist_data = fetch_watchlist_data(symbols, API_KEY)
    if not watchlist_data:
        print("Could not fetch the watchlist data: Exiting.")
        return

    save_successful = save_raw_json(watchlist_data, raw_data_file_name, raw_data_path)
    if not save_successful:
        print("Save failed, skipping load: Exiting.")
        return

    json_file_content = load_raw_json(raw_data_path)
    if not json_file_content:
        print("No json file to fetch: Exiting.")
        return


    symbols_data_list = parse_watchlist_data(json_file_content)
    if not symbols_data_list:
        print("Failed to parse symbols data: Exiting.")
        return

    folder_path = build_date_folder(base_path, today_date)
    if not folder_path:
        print("No folder created: Exiting.")
        return

    archive_daily_files([raw_data_path], folder_path)

    database_file_path = init_database(db_path)
    if not database_file_path:
        print("No database created: Exiting.")
        return

    insert_successful = insert_daily_records(database_file_path, symbols_data_list)
    if not insert_successful:
        print("Insert failed: today's records were not saved. Exiting.")
        return

    rows = fetch_last_n_days(database_file_path, 7)
    if not rows:
        print("Could not fetch last_n days of data: Exiting.")
        return

    reports_folder.mkdir(parents=True, exist_ok=True)
    build_weekly_report(rows, reports_folder)

if __name__ == "__main__":
    main()

